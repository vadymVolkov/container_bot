from datetime import datetime
import openpyxl
import os
import database


def name_generator():
    date = datetime.today()
    date_new = str(date).split('-')
    year = date_new[0]
    month = date_new[1]
    day = date_new[2].split(' ')[0]
    name = year + '_' + month + '_' + day
    return name


def set_data_file_name():
    name = name_generator()
    file_name = name + '.xlsx'
    return file_name


def check_file_in_data_folder(file_name):
    list_files = os.listdir('data/')
    result = list_files.count(file_name)
    if result == 0:
        return False
    else:
        return True


def get_file_data(file_name):
    wb = openpyxl.load_workbook(filename='data/' + file_name)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = sheet.rows
    result = []
    for rows in data:
        temp = []
        for col in rows:
            temp.append(col.value)
        result.append(temp)
    return result


def delete_empty_city_cells_rows_file(file_name):
    wb = openpyxl.load_workbook(filename='data/' + file_name)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = sheet.rows
    result = []
    how_much_delete = 0
    count = 1
    for rows in data:
        temp = []
        for col in rows:
            temp.append(col.value)
        if temp[14] == None or temp[14] == 'Неизвестный':
            sheet.delete_rows(count, 1)
            how_much_delete = how_much_delete + 1
        count = count + 1
    wb.save('data/' + file_name)
    if how_much_delete > 0:
        delete_empty_city_cells_rows_file(file_name)
    else:
       pass


def check_if_city_in_city_table(city_name):
    status = database.search_city(city_name)
    if status:
        return True
    else:
        return False


def change_city_name(file_name):
    wb = openpyxl.load_workbook(filename='data/' + file_name)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = sheet.rows
    for rows in data:
        col_count = 0
        for col in rows:
            col_count = col_count + 1
            if col_count == 15:
                city_name = col.value
                city_name = city_name.lower()
                city_name = city_name.strip()
                try:
                    tmp = city_name.split('.')
                    if len(tmp[0]) < 4:
                        city_name = tmp[1]
                    else:
                        city_name = tmp[0]
                except:
                    pass
                try:
                    tmp = city_name.split(',')
                    if len(tmp[0]) < 4:
                        city_name = tmp[1]
                    else:
                        city_name = tmp[0]
                except:
                    pass
                try:
                    tmp = city_name.split(' ')
                    if len(tmp[0]) < 4:
                        city_name = tmp[1]
                    elif len(tmp[1]) < 4:
                        city_name = tmp[0]
                    else:
                        city_name = tmp[0] + ' ' + tmp[1]
                except:
                    pass
                col.value = city_name
    wb.save('data/' + file_name)


def upload_file_in_database(file_name, database_name):
    data = get_file_data(file_name)
    try:
        for d in data:
            city_name = d[14]
            possible_city_name_list = database.search_city(city_name)
            if possible_city_name_list is None:
                possible_city_name_list = database.find_original_city_by_name(city_name)
            try:
                possible_city_name = possible_city_name_list[1]
            except Exception as e:
                print(e)
                return False
            d[14] = possible_city_name
            database.insert_data_in_table(database_name, d)
        return True
    except Exception as e:
        print(e)
        return False


def get_tables_list_message():
    db_list = database.get_tables_list()
    result = 'Список таблиц:\n'
    for db in db_list:
        result = result + db[1] + '\n'
    return result


def get_table_by_name(name):
    table_from_list = database.get_table_by_name_from_table_list(name)
    return table_from_list


def check_tables_list(message):
    db_list = database.get_tables_list()
    for db in db_list:
        if db[1] == message.text:
            return True
    return False


def check_tables_list_for_client(message):
    db_list = database.get_tables_list_for_client()
    for db in db_list:
        if db[1] == message.text:
            return True
    return False


def create_new_table():
    name = name_generator()
    database.insert_table_name_in_table_list(name)
    try:
        database.create_new_table(name)
        return True
    except Exception as e:
        print(e)
        return False


def check_table_in_db_by_name(name):
    result = database.get_table_by_name_from_table_list(name)
    if result:
        return True
    else:
        return False


def delete_table_by_name(name):
    database.arched_table_by_name(name)
    database.drop_table_by_name(name)
    os.remove('data/' + name + '.xlsx')


def check_if_table_empty(table_name):
    result = database.check_if_table_empty_byt_name(table_name)
    if result[0] == 0:
        return False
    else:
        return True


def check_continue_work_with_table(mess):
    try:
        if mess.text.split('-')[0] == "Продолжить работу с таблицей":
            return True
        else:
            return False
    except Exception as e:
        print(e)
        return False


def delete_empty_recorde_from_table(table_name):
    database.delete_empty_records(table_name)


def set_file_upload_status(status, table_name):
    database.set_file_upload_status_in_table_list(status, table_name)


def set_file_format_status(status, table_name):
    database.set_file_format_status_in_table_list(status, table_name)


def get_city_list_from_file(file_name):
    wb = openpyxl.load_workbook(filename='data/' + file_name)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    city_list = []
    data = sheet.rows
    for rows in data:
        col_count = 0
        for col in rows:
            col_count = col_count + 1
            if col_count == 15:
                city_list.append(col.value)
    return city_list


def get_unique_unknown_citys(file_name):
    citys = get_city_list_from_file(file_name)
    result_step1 = []
    for city in citys:
        a = result_step1.count(city)
        if a == 0:
            result_step1.append(city)
    result_step2 = []
    for city in result_step1:
        city_status = check_if_city_in_city_table(city)
        if city_status:
            pass
        else:
            result_step2.append(city)
    return result_step2


def check_if_original_city_in_city_table(entered_original_city_name):
    city_name = database.find_original_city_by_name(entered_original_city_name)
    if city_name:
        return city_name[1]
    else:
        return False


def add_possible_city_name_to_original(original_city_name_from_city_list, unknown_city_name):
    database.add_possible_city_name_by_original_city_name(original_city_name_from_city_list, unknown_city_name)


def add_new_original_city_name_and_possible(entered_original_city_name, unknown_city_name):
    database.add_new_original_city_name_and_possible_city_name(entered_original_city_name, unknown_city_name)


def set_table_status(status, table_name):
    database.set_table_status_in_table_list(status, table_name)


def get_tables_list_message_for_client():
    db_list = database.get_tables_list_for_client()
    result = 'Список таблиц:\n'
    for db in db_list:
        result = result + db[1] + '\n'
    return result


def get_container_list_by_city_name_in_table(table_name, city_name):
    container_list = database.get_container_list_by_city_name(table_name,city_name)
    return container_list


def set_file_in_db_status(status, table_name):
    database.set_file_in_db_stAtus_in_table_list(status, table_name)

